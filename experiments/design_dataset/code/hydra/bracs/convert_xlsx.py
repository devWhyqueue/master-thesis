"""Convert BRACS.xlsx to a normalised CSV without pandas or the container."""

from __future__ import annotations

import csv
import logging
import pathlib
import re
import sys
from openpyxl import load_workbook  # available on login/compute nodes via system Python

logger = logging.getLogger(__name__)

_FIELDS = ["case_id", "slide_id", "roi_id", "cancer_type", "dataset", "lesion_type"]
_ALIASES = {
    "NORMAL": "N",
    "N": "N",
    "PB": "PB",
    "PATHOLOGICALBENIGN": "PB",
    "UDH": "UDH",
    "USUALDUCTALHYPERPLASIA": "UDH",
    "FEA": "FEA",
    "FLATEPITHELIALATYPIA": "FEA",
    "ADH": "ADH",
    "ATYPICALDUCTALHYPERPLASIA": "ADH",
    "DCIS": "DCIS",
    "DUCTALCARCINOMAINSITU": "DCIS",
    "IC": "IC",
    "INVASIVECARCINOMA": "IC",
}
_LESION = {
    "N": "benign",
    "PB": "benign",
    "UDH": "benign",
    "FEA": "atypical",
    "ADH": "atypical",
    "DCIS": "malignant",
    "IC": "malignant",
}


def _canon(v: object) -> str:
    return re.sub(r"[^a-z0-9]+", "_", str(v).strip().lower()).strip("_")


def _first(names: tuple[str, ...], cols: dict[str, int]) -> int | None:
    """Return the first column index matching any of the candidate names."""
    for n in names:
        if n in cols:
            return cols[n]
    for k, v in cols.items():
        if any(n in k for n in names):
            return v
    return None


def _clean(v: object) -> str:
    s = str(v).strip()
    return "" if s.lower() == "nan" else pathlib.Path(s).stem


def _find_columns(
    header_cells: list,
) -> tuple[int | None, int | None, int | None, int | None]:
    """Return (roi_idx, slide_idx, case_idx, label_idx) column positions."""
    cols = {_canon(c.value): i for i, c in enumerate(header_cells)}
    return (
        _first(("roi", "roi_id", "roi_filename", "roi_name", "image"), cols),
        _first(("wsi", "slide", "slide_id", "wsi_filename"), cols),
        _first(("patient", "patient_id", "case", "case_id"), cols),
        _first(
            ("subtype", "lesion_subtype", "label", "class", "diagnosis", "category"),
            cols,
        ),
    )


def _load_sheet_rows(xlsx: pathlib.Path) -> tuple[list, list]:
    """Return header cells and data rows from the largest sheet."""
    wb = load_workbook(xlsx, read_only=True)
    largest = max(wb.worksheets, key=lambda ws: ws.max_row)
    rows = list(largest.rows)
    return list(rows[0]), rows[1:]


def _make_row(vals: list, ri: int, si: int, ci: int, li: int) -> dict | None:
    """Build a normalised CSV row dict, or None to skip."""
    if all(v is None for v in vals):
        return None
    label = _ALIASES.get(re.sub(r"[^A-Z0-9]+", "", str(vals[li]).strip().upper()))
    roi, slide, case = _clean(vals[ri]), _clean(vals[si]), _clean(vals[ci])
    if not (label and roi and slide and case):
        return None
    return {
        "case_id": case,
        "slide_id": slide,
        "roi_id": roi,
        "cancer_type": label,
        "dataset": "bracs",
        "lesion_type": _LESION[label],
    }


def _write_csv(output_csv: pathlib.Path, rows: list) -> None:
    """Write normalised rows to CSV."""
    output_csv.parent.mkdir(parents=True, exist_ok=True)
    with open(output_csv, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=_FIELDS)
        w.writeheader()
        w.writerows(rows)


def convert(data_root: pathlib.Path, output_csv: pathlib.Path) -> int:
    """Convert BRACS.xlsx to a normalised CSV and return the row count."""
    xlsx = next(data_root.rglob("BRACS.xlsx"))
    header_cells, data_rows = _load_sheet_rows(xlsx)
    ri, si, ci, li = _find_columns(header_cells)
    if ri is None or si is None or ci is None or li is None:
        raise ValueError(f"Could not identify required columns in {xlsx}")
    out = [
        r
        for row in data_rows
        if (r := _make_row([c.value for c in row], ri, si, ci, li))
    ]
    _write_csv(output_csv, out)
    return len(out)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO, format="%(message)s")
    data_root, output_csv = pathlib.Path(sys.argv[1]), pathlib.Path(sys.argv[2])
    logger.info("wrote %d rows to %s", convert(data_root, output_csv), output_csv)
